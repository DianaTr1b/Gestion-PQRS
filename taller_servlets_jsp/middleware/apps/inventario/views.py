from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q , Count
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Elemento, MovimientoInventario, Categoria, DetalleMovimiento,NombreElemento
from .forms import (
    ElementoForm, MovimientoInventarioForm, DetalleMovimientoFormSet,
    BusquedaElementoForm, BusquedaMovimientoForm
)
from django.http import HttpResponse
from apps.usuarios.models import PerfilUsuario
from apps.usuarios.services.directorio_service import sincronizar_directorio
from datetime import datetime, timedelta
from django.db import transaction
from django.urls import reverse
from apps.usuarios.permissions import(
    tecnico_required,
    es_tecnico,
    es_bodega,
    es_colaborador, 
    login_required_custom,
    colaborador_restringido,
    solo_tecnico,
)
from django.views.decorators.csrf import csrf_exempt
from .email_utils import enviar_notificacion_firma
from .email_utils import enviar_notificacion_movimiento, enviar_recordatorio_firma
from .services import guardar_firma_acta
from django.template.loader import get_template
from xhtml2pdf import pisa
import os
from django.conf import settings
#===========================================================

def asignar_documento(request, doc_id, user_id):
    # ... lógica para obtener documento y usuario ...
    # documento.usuario_asignado = usuario
    # documento.save()

    # Llamamos a la función de notificación
    enviar_notificacion_firma(usuario, documento)


# ===================== APIs PARA AJAX =====================

# En tu archivo views.py

@solo_tecnico
@require_http_methods(["GET"])
def obtener_elementos_usuario(request, usuario_id):
    """API para obtener elementos asignados a un usuario o bodega"""
    try:
        from apps.usuarios.models import PerfilUsuario
        usuario = PerfilUsuario.objects.select_related('rol').get(id=usuario_id)

        # SOLUCIÓN: Diferenciar la consulta si el Origen es Bodega o Colaborador
        if usuario.rol and usuario.rol.nombre == 'Bodega':
            elementos = Elemento.objects.filter(
                bodega_actual_id=usuario_id,
                estado__in=['Disponible', 'Mantenimiento'],
                usuario_actual=None
            ).exclude(detalles_movimiento__movimiento__estado_movimiento='Pendiente')
        else:
            elementos = Elemento.objects.filter(
                usuario_actual_id=usuario_id,
                estado='Asignado'
            ).exclude(detalles_movimiento__movimiento__estado_movimiento='Pendiente')

        elementos_list = []
        for elemento in elementos:
            elementos_list.append({
                'id': elemento.id,
                'nombre': elemento.nombre,
                'marca': elemento.marca,
                'modelo': elemento.modelo,
                'serial': elemento.serial or 'sin serial',
                'categoria': elemento.categoria.nombre,
                'bodega': str(elemento.bodega_actual),
                'estado': elemento.get_estado_display() if hasattr(elemento, 'get_estado_display') else elemento.estado,
            })
        return JsonResponse({
            'elementos': elementos_list,
            'total': len(elementos_list),
            'usuario_id': usuario_id
        })

    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'elementos': [],
            'total': 0,
        }, status=200)

# ===================== DASHBOARD =====================

@login_required_custom
def dashboard(request):
    """Vista principal del dashboard con estadísticas"""
    perfil = request.user.perfil

    # ── Dashboard Colaborador ─────────────────────────────────────────────
    if es_colaborador(request.user):

        elementos_propios = Elemento.objects.filter(
            usuario_actual=perfil
        ).select_related('categoria', 'bodega_actual')

        movimientos_propios = MovimientoInventario.objects.filter(
            Q(usuario_destino=perfil) | Q(usuario_autoriza=perfil)
        ).select_related(
            'usuario_registro', 'usuario_origen', 'usuario_destino'
        ).order_by('-fecha_movimiento')[:10]

        # CORRECCIÓN: Filtramos los que faltan por firmar específicamente por el usuario actual
        pendientes_firma = MovimientoInventario.objects.filter(
            (Q(usuario_destino=perfil) & (Q(firma_recibe__isnull=True) | Q(firma_recibe=''))) |
            (Q(usuario_autoriza=perfil) & (Q(firma_autoriza__isnull=True) | Q(firma_autoriza='')))
        ).distinct()

        # Actualizamos también la sección de movimientos recientes para este usuario
        movimientos_recientes = MovimientoInventario.objects.filter(
            Q(usuario_origen=perfil) | Q(usuario_destino=perfil) | Q(usuario_autoriza=perfil)
        ).select_related('usuario_registro', 'usuario_origen', 'usuario_destino').order_by('-fecha_movimiento')[:10]

        return render(request, 'apps/inventario/dashboard_colaborador.html', {
            'elementos_propios': elementos_propios,
            'total_elementos_propios': elementos_propios.count(),
            'movimientos_propios': movimientos_recientes, # ✅ Ahora usamos la lista filtrada
            'pendientes_firma': pendientes_firma,
            'total_pendientes': pendientes_firma.count(),
        })

    # ── Dashboard Bodega ──────────────────────────────────────────────────
    if es_bodega(request.user):
        elementos_bodega = Elemento.objects.filter(bodega_actual=perfil)
        total_elementos      = elementos_bodega.count()
        elementos_disponibles  = elementos_bodega.filter(estado='Disponible').count()
        elementos_asignados    = elementos_bodega.filter(estado='Asignado').count()
        elementos_mantenimiento = elementos_bodega.filter(estado='Mantenimiento').count()
        elementos_baja         = elementos_bodega.filter(estado='Baja').count()

        elementos_por_categoria = Categoria.objects.filter(
            elementos__bodega_actual=perfil
        ).annotate(total=Count('elementos')).order_by('-total')[:5]

        elementos_por_bodega = PerfilUsuario.objects.filter(
            id=perfil.id
        ).annotate(total=Count('elementos_en_bodega'))

        ultimos_movimientos = MovimientoInventario.objects.filter(
            Q(usuario_origen=perfil) | Q(usuario_destino=perfil) | Q(usuario_autoriza=perfil)
        ).select_related(
            'usuario_registro', 'usuario_origen', 'usuario_destino'
        ).order_by('-fecha_movimiento')[:10]

        # NUEVO: Lógica de firmas para Bodega (como Elabora, Recibe o Autoriza)
        pendientes_firma = MovimientoInventario.objects.filter(
            (Q(usuario_destino=perfil) & (Q(firma_recibe__isnull=True) | Q(firma_recibe=''))) |
            (Q(usuario_registro=perfil) & (Q(firma_elabora__isnull=True) | Q(firma_elabora=''))) |
            (Q(usuario_autoriza=perfil) & (Q(firma_autoriza__isnull=True) | Q(firma_autoriza='')))
        ).distinct()

        fecha_limite = datetime.now().date() + timedelta(days=30)
        
        # 1. Equipos Físicos (Excluye Suscripciones)
        elementos_garantia_proxima = Elemento.objects.filter( # Si es bodega usa 'elementos_bodega.filter'
            garantia_hasta__lte=fecha_limite,
            garantia_hasta__gte=datetime.now().date()
        ).exclude(condicion='Suscripcion').order_by('garantia_hasta')[:5]
        
        # 2. Suscripciones y Licencias
        suscripciones_proximas = Elemento.objects.filter( # Si es bodega usa 'elementos_bodega.filter'
            condicion='Suscripcion',
            garantia_hasta__lte=fecha_limite,
            garantia_hasta__gte=datetime.now().date()
        ).order_by('garantia_hasta')[:5]

        return render(request, 'apps/inventario/dashboard.html', {
            'total_elementos': total_elementos,
            'elementos_disponibles': elementos_disponibles,
            'elementos_asignados': elementos_asignados,
            'elementos_mantenimiento': elementos_mantenimiento,
            'elementos_baja': elementos_baja,
            'elementos_por_categoria': elementos_por_categoria,
            'elementos_por_bodega': elementos_por_bodega,
            'ultimos_movimientos': ultimos_movimientos,
            'elementos_garantia_proxima': elementos_garantia_proxima,
            'suscripciones_proximas': suscripciones_proximas,
            'pendientes_firma': pendientes_firma,
            'total_pendientes': pendientes_firma.count(),
        })

    # ── Dashboard Técnico ─────────────────────────────────────────────────
    total_elementos = Elemento.objects.count()
    elementos_disponibles = Elemento.objects.filter(estado='Disponible').count()
    elementos_asignados = Elemento.objects.filter(estado='Asignado').count()
    elementos_mantenimiento = Elemento.objects.filter(estado='Mantenimiento').count()
    elementos_baja = Elemento.objects.filter(estado='Baja').count()

    elementos_por_categoria = Categoria.objects.annotate(
        total=Count('elementos')
    ).order_by('-total')[:5]

    elementos_por_bodega = PerfilUsuario.objects.filter(
        rol__nombre='Bodega'
    ).annotate(total=Count('elementos_en_bodega')).order_by('-total')

    ultimos_movimientos = MovimientoInventario.objects.all().select_related(
        'usuario_registro', 'usuario_origen', 'usuario_destino'
    ).order_by('-fecha_movimiento')[:10]

    elementos_propios = Elemento.objects.filter(
        usuario_actual=perfil
    ).select_related('categoria', 'bodega_actual')
    
    movimientos_propios = MovimientoInventario.objects.filter(
        Q(usuario_origen=perfil) | Q(usuario_destino=perfil) | Q(usuario_autoriza=perfil)
    ).select_related('usuario_registro', 'usuario_origen', 'usuario_destino').order_by('-fecha_movimiento')[:10]

    pendientes_firma = MovimientoInventario.objects.filter(
        (Q(usuario_destino=perfil) & (Q(firma_recibe__isnull=True) | Q(firma_recibe=''))) |
        (Q(usuario_registro=perfil) & (Q(firma_elabora__isnull=True) | Q(firma_elabora=''))) |
        (Q(usuario_autoriza=perfil) & (Q(firma_autoriza__isnull=True) | Q(firma_autoriza='')))
    ).distinct()

    fecha_limite = datetime.now().date() + timedelta(days=30)
    # 1. Equipos Físicos (Excluye Suscripciones)
    elementos_garantia_proxima = Elemento.objects.filter( # Si es bodega usa 'elementos_bodega.filter'
        garantia_hasta__lte=fecha_limite,
        garantia_hasta__gte=datetime.now().date()
    ).exclude(condicion='Suscripcion').order_by('garantia_hasta')[:5]
    
    # 2. Suscripciones y Licencias
    suscripciones_proximas = Elemento.objects.filter( # Si es bodega usa 'elementos_bodega.filter'
        condicion='Suscripcion',
        garantia_hasta__lte=fecha_limite,
        garantia_hasta__gte=datetime.now().date()
    ).order_by('garantia_hasta')[:5]

    return render(request, 'apps/inventario/dashboard_colaborador.html', {
        'total_elementos': total_elementos,
        'elementos_disponibles': elementos_disponibles,
        'elementos_asignados': elementos_asignados,
        'elementos_mantenimiento': elementos_mantenimiento,
        'elementos_baja': elementos_baja,
        'elementos_por_categoria': elementos_por_categoria,
        'elementos_por_bodega': elementos_por_bodega,
        'ultimos_movimientos': ultimos_movimientos,
        'elementos_garantia_proxima': elementos_garantia_proxima,
        'suscripciones_proximas': suscripciones_proximas,

        'elementos_propios': elementos_propios,
        'total_elementos_propios': elementos_propios.count(),
        'movimientos_propios': movimientos_propios,

        'pendientes_firma': pendientes_firma,
        'total_pendientes': pendientes_firma.count(),
    })


# ===================== ELEMENTOS - CRUD =====================

@login_required_custom
@colaborador_restringido
def elementos_lista(request):
    """Lista todos los elementos con filtros y paginación"""
    user = request.user

    if es_colaborador(user):
        elementos = Elemento.objects.filter(usuario_actual=user.perfil)
    elif es_bodega(user):
        elementos =Elemento.objects.filter(bodega_actual=user.perfil)
    else:
        elementos = Elemento.objects.all()
    
    # Filtros
    estado_filtro = request.GET.get('estado')
    categoria_filtro = request.GET.get('categoria')
    bodega_filtro = request.GET.get('bodega')
    busqueda = request.GET.get('q')
    orden = request.GET.get('orden', '-fecha_registro')
    id_filtro = request.GET.get('id_elemento','')

    if estado_filtro:
        elementos = elementos.filter(estado=estado_filtro)
    if categoria_filtro:
        elementos = elementos.filter(categoria_id=categoria_filtro)
    if bodega_filtro:
        elementos = elementos.filter(bodega_actual_id=bodega_filtro)
    if id_filtro:
        elementos = elementos.filter(id=id_filtro)   
    
    if busqueda:
        if busqueda.isdigit():
            elementos = elementos.filter(
                Q(id=int(busqueda)) |
                Q(nombre_elemento__nombre__icontains=busqueda) |
                Q(marca__icontains=busqueda) |
                Q(modelo__icontains=busqueda) |
                Q(serial__icontains=busqueda)
            )
        else:
            elementos = elementos.filter(
                Q(nombre_elemento__nombre__icontains=busqueda) |
                Q(marca__icontains=busqueda) |
                Q(modelo__icontains=busqueda) |
                Q(serial__icontains=busqueda)
            )
    
    elementos = elementos.order_by(orden)
    
    paginator = Paginator(elementos, 10)
    page = request.GET.get('page', 1)
    
    try:
        elementos_paginados = paginator.page(page)
    except PageNotAnInteger:
        elementos_paginados = paginator.page(1)
    except EmptyPage:
        elementos_paginados = paginator.page(paginator.num_pages)
    
    categorias = Categoria.objects.filter(estado='Activo').order_by('nombre')
    bodegas = PerfilUsuario.objects.filter(rol__nombre = 'Bodega', estado='Activo').order_by('user__first_name')
    
    context = {
        'elementos': elementos_paginados,
        'categorias': categorias,
        'bodegas': bodegas,
        'estado_filtro': estado_filtro,
        'categoria_filtro': categoria_filtro,
        'bodega_filtro': bodega_filtro,
        'busqueda': busqueda,
        'orden': orden,
        'id_filtro': id_filtro,
    }
    
    return render(request, 'apps/inventario/elementos/lista.html', context)


@solo_tecnico
@colaborador_restringido
def elemento_crear(request):
    perfil =request.user.perfil
    if request.method == 'POST':
        form = ElementoForm(request.POST, request.FILES, usuario=perfil)
        if form.is_valid():
            elemento =form.save(commit=False)
            elemento.usuario_registro = perfil
            elemento.save()
            messages.success(request, f'Elemento creado exitosamente.')
            return redirect('inventario:elementos_lista')
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    messages.error(request, f'{label}: {error}')
    else:
        form = ElementoForm(usuario=perfil)
    
    context = {
        'form': form,
        'titulo': 'Crear Elemento',
        'accion': 'Guardar'
    }
    return render(request, 'apps/inventario/elementos/formulario.html', context)


@solo_tecnico
@colaborador_restringido
def elemento_editar(request, elemento_id):
    elemento = get_object_or_404(Elemento, id=elemento_id)
    perfil = request.user.perfil
    if request.method == 'POST':
        form = ElementoForm(request.POST, instance=elemento, usuario=perfil)
        if form.is_valid():
            elemento = form.save(commit=False)
            elemento.usuario_registro = perfil
            elemento.save()
            messages.success(request, f'Elemento "{elemento.nombre}" actualizado exitosamente.')
            return redirect('inventario:elemento_detalle', elemento_id=elemento.id)
        
        else:
            for field, errors in form.errors.items():
                label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    messages.error(request, f'{label}: {error}')
    else:
        form = ElementoForm(instance=elemento, usuario=perfil)
    
    context = {
        'form': form,
        'elemento': elemento,
        'titulo': f'Editar: {elemento.nombre}',
        'accion': 'Actualizar'
    }
    
    return render(request, 'apps/inventario/elementos/formulario.html', context)


@login_required_custom
def elemento_detalle(request, elemento_id):
    elemento = get_object_or_404(
        Elemento.objects.select_related(
            'categoria', 'bodega_actual', 'usuario_actual', 'usuario_registro'
        ),
        id=elemento_id
    )
    
    historial = DetalleMovimiento.objects.filter(
        elemento=elemento
    ).select_related(
        'movimiento__usuario_registro',
        'movimiento__usuario_origen',
        'movimiento__usuario_destino'
    ).order_by('-movimiento__fecha_movimiento')
    
    try:
        from .models import HistorialUsuario
        historial_usuarios = HistorialUsuario.objects.filter(
            elemento=elemento
        ).select_related(
            'usuario_anterior',
            'usuario_nuevo',
            'movimiento'
        ).order_by('-fecha_cambio')
    except:
        historial_usuarios = None

    url_retorno = request.META.get('HTTP_REFERER')    
    
    if not url_retorno:
        url_retorno = reverse('inventario:elementos_lista')
    context = {
        'elemento': elemento,
        'historial': historial,
        'historial_usuarios': historial_usuarios,
        # 'query_parms': request.GET.urlencode(),
        'url_retorno': url_retorno,
    }
    
    return render(request, 'apps/inventario/elementos/detalle.html', context)


@solo_tecnico
def elemento_eliminar(request, elemento_id):
    elemento = get_object_or_404(Elemento, id=elemento_id)
    tiene_movimientos = DetalleMovimiento.objects.filter(elemento=elemento).exists()
    
    if request.method == 'POST':
        if tiene_movimientos:
            messages.error(
                request, 
                f'No se puede eliminar "{elemento.nombre}" porque tiene movimientos asociados. '
                'Considera marcarlo como "Baja" en su lugar.'
            )
            return redirect('inventario:elemento_detalle', elemento_id=elemento.id)
        
        nombre_elemento = elemento.nombre
        elemento.delete()
        messages.success(request, f'Elemento "{nombre_elemento}" eliminado exitosamente.')
        return redirect('inventario:elementos_lista')
    
    context = {
        'elemento': elemento,
        'tiene_movimientos': tiene_movimientos,
    }

    return render(request, 'apps/inventario/elementos/eliminar.html', context)


# ===================== MOVIMIENTOS - CRUD =====================

@login_required_custom
@colaborador_restringido
def movimientos_lista(request):
    """Lista todos los movimientos con filtros y paginación"""
    user = request.user

    if es_colaborador(user):
    # CORRECCIÓN: Permitir ver movimientos al usuario autorizador
        movimientos = MovimientoInventario.objects.filter(
            Q(usuario_origen=user.perfil) | Q(usuario_destino=user.perfil) | Q(usuario_autoriza=user.perfil)
        ).select_related()
    elif es_bodega(user):
        movimientos = MovimientoInventario.objects.filter(
            Q(usuario_origen=user.perfil) | Q(usuario_destino=user.perfil) | Q(usuario_autoriza=user.perfil)
        ).select_related()
    else:
        movimientos = MovimientoInventario.objects.all().select_related()


    tipo_filtro = request.GET.get('tipo')
    estado_filtro = request.GET.get('estado')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    orden = request.GET.get('orden', '-fecha_movimiento')
    
    if tipo_filtro:
        movimientos = movimientos.filter(tipo_movimiento=tipo_filtro)
    if estado_filtro:
        movimientos = movimientos.filter(estado_movimiento=estado_filtro)
    if fecha_desde:
        movimientos = movimientos.filter(fecha_movimiento__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha_movimiento__lte=fecha_hasta)
    
    movimientos = movimientos.order_by(orden)
    
    paginator = Paginator(movimientos, 10)
    page = request.GET.get('page', 1)
    
    try:
        movimientos_paginados = paginator.page(page)
    except PageNotAnInteger:
        movimientos_paginados = paginator.page(1)
    except EmptyPage:
        movimientos_paginados = paginator.page(paginator.num_pages)
    
    context = {
        'movimientos': movimientos_paginados,
        'tipo_filtro': tipo_filtro,
        'estado_filtro': estado_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'orden': orden,
    }

    return render(request, 'apps/inventario/movimientos/lista.html', context)


@solo_tecnico
@colaborador_restringido
def movimiento_crear(request):
    perfil = request.user.perfil
    if request.method == 'POST':
        form = MovimientoInventarioForm(request.POST, request.FILES, usuario=perfil)
        if form.is_valid():
            try: 
                with transaction.atomic():
                    movimiento = form.save(commit=False)
                    movimiento.usuario_registro = perfil
                    movimiento.save()
                    formset = DetalleMovimientoFormSet(
                        request.POST,
                        instance=movimiento,
                        form_kwargs={
                            'tipo_movimiento': movimiento.tipo_movimiento,
                            'usuario_origen': movimiento.usuario_origen,
                        }
                    )       
            
                    if formset.is_valid():
                        detalles_validos= [f for f in formset if f.cleaned_data.get('elemento')]
                        if not detalles_validos:
                            messages.error(request,'Debe agregar al menos un elemento al movimiento.')
                            raise Exception('Sin elementos')
                        formset.save()

                        if movimiento.necesita_notificacion_firma:
                            try:
                                # Llamamos a la utilidad que ya tienes en email_utils.py
                                enviar_notificacion_movimiento(movimiento)
                            except Exception as email_err:
                                # Si el correo falla, lo registramos pero no interrumpimos el flujo
                                print(f"Error no crítico al enviar correo: {email_err}")
                        # -----------------------------------------------

                        messages.success(
                            request, 
                            f'Movimiento #{movimiento.id} de tipo "{movimiento.get_tipo_movimiento_display()}" creado exitosamente.'
                        )
                        return redirect('inventario:movimientos_lista')
                    else:
                        for i, f in enumerate(formset):
                            if f.errors:
                                print(f"Form {i} errors:{f.errors}")
                            else:
                                print(f"Form {i} sin errores, data: {f.data}")    
                        print(f"Formset non_form_errors:{formset.non_form_errors()}")
                        raise Exception('Formset Inválido')        
                    
            except Exception as e:
                # El error se imprime en consola y la transacción hace rollback
                print(f"EXCEPCION EN CREACIÓN: {e}")
                messages.error(request, 'Ocurrió un error al procesar el movimiento. Verifique los datos.')
                
        else:
            formset = DetalleMovimientoFormSet(request.POST)
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = MovimientoInventarioForm(usuario=perfil)
        formset = DetalleMovimientoFormSet()

    sincronizar_directorio()
    
    # Contexto para el renderizado
    usuarios = PerfilUsuario.objects.filter(estado='Activo').select_related('user').order_by('-id')
    elementos_disponibles = Elemento.objects.filter(estado='Disponible').select_related('categoria', 'bodega_actual')
    
    context = {
        'form': form,
        'formset': formset,
        'usuarios': usuarios,
        'elementos_disponibles': elementos_disponibles,
    }
    return render(request, 'apps/inventario/movimientos/crear.html', context)

# ... código anterior en views.py ...

@solo_tecnico
@colaborador_restringido
def movimiento_editar(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    perfil = request.user.perfil

    # REGLA MEJORADA: Bloquear si está firmado totalmente O si fue anulado
    if movimiento.es_firmado_total or movimiento.estado_movimiento == 'Anulado':
        messages.error(request, 'El documento está bloqueado porque ya fue completado o ha sido anulado.')
        return redirect('inventario:movimiento_detalle', movimiento_id=movimiento.id)
    
    if request.method == 'POST':
        # 1. Cargamos el formulario principal con los datos enviados, sobre la instancia actual
        form = MovimientoInventarioForm(request.POST, request.FILES, instance=movimiento, usuario=perfil)
        
        # 2. Cargamos el FormSet (los elementos) sobre la misma instancia
        formset = DetalleMovimientoFormSet(
            request.POST,
            instance=movimiento,
            form_kwargs={
                'tipo_movimiento': movimiento.tipo_movimiento,
                'usuario_origen': request.POST.get('usuario_origen', movimiento.usuario_origen_id),
            }
        )
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    # Guardamos los cambios del encabezado
                    form.save()
                    
                    # Validamos que no intenten borrar todos los elementos
                    detalles_validos = [f for f in formset if f.cleaned_data and not f.cleaned_data.get('DELETE', False)]
                    if not detalles_validos:
                        messages.error(request, 'El documento debe mantener al menos un elemento asociado.')
                        raise Exception('Sin elementos en la edición')
                        
                    # Guardamos los cambios en los elementos
                    formset.save()
                    
                    messages.success(request, f'El movimiento #{movimiento.id} ha sido actualizado con éxito.')
                    return redirect('inventario:movimiento_detalle', movimiento_id=movimiento.id)
                    
            except Exception as e:
                print(f"Error en edición: {e}")
                messages.error(request, 'Error al guardar los cambios. Por favor, verifica la información.')
        else:
            messages.error(request, 'Corrige los errores indicados en el formulario.')
            
    else:
        # Petición GET: Pintamos el formulario con la información actual de la base de datos
        form = MovimientoInventarioForm(instance=movimiento, usuario=perfil)
        formset = DetalleMovimientoFormSet(
            instance=movimiento,
            form_kwargs={
                'tipo_movimiento': movimiento.tipo_movimiento,
                'usuario_origen': movimiento.usuario_origen,
            }
        )

    usuarios = PerfilUsuario.objects.filter(estado='Activo').select_related('user').order_by('-id')
    elementos_disponibles = Elemento.objects.filter(estado='Disponible').select_related('categoria', 'bodega_actual')
    
    context = {
        'form': form,
        'formset': formset,
        'movimiento': movimiento,
        'usuarios': usuarios,
        'elementos_disponibles': elementos_disponibles,
    }
    
    return render(request, 'apps/inventario/movimientos/editar.html', context)

@solo_tecnico
def movimiento_anular(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    
    if request.method == 'POST':
        if movimiento.estado_movimiento == 'Pendiente':
            movimiento.estado_movimiento = 'Anulado'
            movimiento.save()
            messages.success(request, f'El movimiento #{movimiento.id} ha sido anulado. Los equipos fueron liberados.')
        else:
            messages.error(request, 'Solo se pueden anular movimientos en estado Pendiente.')
            
    return redirect('inventario:movimiento_detalle', movimiento_id=movimiento.id)

@login_required_custom
def movimiento_detalle(request, movimiento_id):
    """Ver detalle de un movimiento"""
    movimiento = get_object_or_404(
        MovimientoInventario.objects.select_related(
            'usuario_registro', 'usuario_origen', 'usuario_destino'
        ),
        id=movimiento_id
    )

    perfil = request.user.perfil

    # CORRECCIÓN: Agregar permisos para que el autorizador no sea bloqueado
    es_tecnico_usuario = es_tecnico(request.user)
    involucrado = (
        movimiento.usuario_destino == perfil or
        movimiento.usuario_origen == perfil or
        movimiento.usuario_registro == perfil or
        movimiento.usuario_autoriza == perfil  # AGREGADO AQUÍ
    )
    if not es_tecnico_usuario and not involucrado:
        messages.error(request, 'No tienes permiso para ver este movimiento.')
        return redirect('inventario:mis_documentos')
    
    detalles = movimiento.detalles.select_related('elemento__categoria')
    
    from django.db.models import Sum
    cantidad_total = detalles.aggregate(total=Sum('cantidad'))['total'] or 0
    
    context = {
        'movimiento': movimiento,
        'detalles': detalles,
        'cantidad_total': cantidad_total,
    }

    return render(request, 'apps/inventario/movimientos/detalle.html', context)


#=====================================================

@login_required_custom
@require_http_methods(["GET"])
def obtener_nombres_elementos(request, categoria_id):
    try:
        nombres = NombreElemento.objects.filter(
            categoria_id=categoria_id,
            activo=True
        ).values('id', 'nombre')
        
        nombres_list = []
        for nombre in nombres:
            obj = NombreElemento.objects.get(id=nombre['id'])
            configuracion = {
                'serial': obj.requiere_serial,
                'imei': obj.requiere_imei,
                'imei2': obj.requiere_imei2,
                'color': obj.requiere_color,
                'marca': obj.requiere_marca,
                'modelo': obj.requiere_modelo,
                'operador': obj.requiere_operador,
                'numero': obj.requiere_numero,
                'capacidad': obj.requiere_capacidad,
                'tipo': obj.requiere_tipo,
                'caracteristica': obj.requiere_caracteristica,
                'puertos': obj.requiere_puertos,
                'mac': obj.requiere_mac,
            }
            
            nombres_list.append({
                'id': nombre['id'],
                'nombre': nombre['nombre'],
                'configuracion': configuracion
            })
        
        return JsonResponse({
            'nombres': nombres_list,
            'categoria_id': categoria_id
        })
    
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'nombres': []
        }, status=500)


# ===================== REPORTES =====================

@login_required_custom
@colaborador_restringido
def reportes(request):
    total_elementos = Elemento.objects.count()
    por_estado = {
        'disponibles': Elemento.objects.filter(estado='Disponible').count(),
        'asignados': Elemento.objects.filter(estado='Asignado').count(),
        'mantenimiento': Elemento.objects.filter(estado='Mantenimiento').count(),
        'baja': Elemento.objects.filter(estado='Baja').count(),
    }
    
    por_categoria = Categoria.objects.annotate(
        total=Count('elementos')
    ).order_by('-total')
    
    fecha_inicio = datetime.now().date() - timedelta(days=30)
    movimientos_mes = MovimientoInventario.objects.filter(
        fecha_movimiento__gte=fecha_inicio
    ).count()
    
    context = {
        'total_elementos': total_elementos,
        'por_estado': por_estado,
        'por_categoria': por_categoria,
        'movimientos_mes': movimientos_mes,
    }

    return render(request, 'apps/inventario/reportes/index.html', context)

# ===================================Nueva API==================================
@solo_tecnico
@require_http_methods(["GET"])
def obtener_elementos_asignados(request):
    try:
        elementos = Elemento.objects.filter(
            estado='Asignado'
        ).select_related('categoria', 'nombre_elemento')

        elementos_list = [{
            'id': e.id,
            'nombre': e.nombre,
            'marca': e.marca,
            'modelo': e.modelo,
        } for e in elementos]

        return JsonResponse({'elementos': elementos_list, 'total': len(elementos_list)})

    except Exception as e:
        return JsonResponse({'error': str(e), 'elementos': []}, status=200)
    
#====================================APIs Usuarios====================================    

@solo_tecnico
def api_usuarios_bodega(request):
    sincronizar_directorio()
    usuarios = PerfilUsuario.objects.filter(
        rol__nombre='Bodega', estado='Activo'
    ).select_related('user')
    data = [{'id': u.id, 'nombre': u.nombre} for u in usuarios]
    return JsonResponse({'usuarios': data})

@solo_tecnico
def api_usuarios_activos(request):
    sincronizar_directorio()
    usuarios = PerfilUsuario.objects.filter(
        estado='Activo'
    ).select_related('user')
    data = [{'id': u.id, 'nombre': u.nombre} for u in usuarios]
    return JsonResponse({'usuarios': data})

@solo_tecnico
def api_usuarios_por_rol(request, rol):
    sincronizar_directorio()
    usuarios = PerfilUsuario.objects.filter(
        rol__nombre=rol, estado='Activo'
    ).select_related('user')
    data = [{'id': u.id, 'nombre': u.nombre} for u in usuarios]
    return JsonResponse({'usuarios': data})

@solo_tecnico
def api_usuarios_excepto_rol(request, rol):
    sincronizar_directorio()
    usuarios = PerfilUsuario.objects.filter(
        estado='Activo'
    ).exclude(rol__nombre=rol).select_related('user')
    data = [{'id': u.id, 'nombre': u.nombre} for u in usuarios]
    return JsonResponse({'usuarios': data})

#=======================API ELEMENTOS BODEGA=========================
@solo_tecnico
def api_elementos_bodega(request, usuario_id):
    elementos = Elemento.objects.filter(
        bodega_actual_id=usuario_id,
        estado__in=['Disponible','Mantenimiento'],
        usuario_actual=None
    ).select_related('categoria')
    data = [{'id': e.id, 'nombre': e.nombre, 'marca': e.marca, 'modelo': e.modelo,'estado':e.estado} for e in elementos]
    return JsonResponse({'elementos': data})

#=====================API PARA MANTENIMIENTO=========================
@solo_tecnico
def api_elementos_para_mantenimiento(request, usuario_id):
    try:
        from apps.usuarios.models import PerfilUsuario
        usuario = PerfilUsuario.objects.select_related('rol').get(id=usuario_id)
        
        if usuario.rol and usuario.rol.nombre == 'Bodega':
            elementos = Elemento.objects.filter(
                bodega_actual_id=usuario_id,
                usuario_actual=None
            ).exclude(estado='Baja')
        else:
            elementos = Elemento.objects.filter(
                usuario_actual_id=usuario_id,
                estado='Asignado'
            )
        
        data = [{'id': e.id, 'nombre': e.nombre, 'marca': e.marca, 'modelo': e.modelo, 'estado': e.estado} for e in elementos]
        return JsonResponse({'elementos': data})
    except Exception as e:
        return JsonResponse({'elementos': [], 'error': str(e)})
    
#===================================================================    
@solo_tecnico
def usuarios_con_elementos(request):
    user = request.user

    if es_bodega(user):
        usuarios = PerfilUsuario.objects.filter(
            id=user.perfil.id
        ).prefetch_related('elementos_asignados__categoria', 'elementos_asignados__nombre_elemento')
    else:
        usuarios = PerfilUsuario.objects.filter(
            elementos_asignados__isnull=False
        ).distinct().prefetch_related(
            'elementos_asignados__categoria',
            'elementos_asignados__nombre_elemento',
            'rol'
        )

    context = {
        'usuarios': usuarios,
    }
    return render(request, 'apps/inventario/reportes/usuarios_elementos.html', context)

#====================================REPORTES===============================

@login_required_custom
@colaborador_restringido
def reporte_por_bodega(request):
    from django.http import HttpResponse
    exportar = request.GET.get('exportar')
    bodega_filtro = request.GET.get('bodega', '')
    estado_filtro = request.GET.get('estado', '')
    categoria_filtro = request.GET.get('categoria', '')
    nombre_filtro = request.GET.get('nombre', '')
    fecha_compra_desde = request.GET.get('fecha_desde', '')
    fecha_compra_hasta = request.GET.get('fecha_hasta', '')
    id_filtro=request.GET.get('id_elemento','')
    
    elementos = Elemento.objects.filter(
        bodega_actual__isnull=False
    ).select_related('categoria', 'bodega_actual', 'usuario_actual')

    user = request.user

    if es_bodega(user):
        perfil_bodega = user.perfil
        if perfil_bodega.id == 1: 
            pass
        else:
            elementos = elementos.filter(bodega_actual=perfil_bodega)
            bodega_filtro = str(perfil_bodega.id)

    if bodega_filtro:
        elementos = elementos.filter(bodega_actual_id=bodega_filtro)
    if estado_filtro:
        elementos = elementos.filter(estado=estado_filtro)
    if categoria_filtro:
        elementos = elementos.filter(categoria_id=categoria_filtro)
    if nombre_filtro:
        elementos = elementos.filter(nombre_elemento_id=nombre_filtro)
    if fecha_compra_desde:
        elementos = elementos.filter(fecha_compra__gte=fecha_compra_desde)
    if fecha_compra_hasta:
        elementos = elementos.filter(fecha_compra__lte=fecha_compra_hasta)
    if id_filtro:
        elementos = elementos.filter(id=id_filtro)

    base_resumen = Elemento.objects.filter(bodega_actual__isnull=False)
    if es_bodega(user) and perfil_bodega.id != 1:  
        base_resumen = base_resumen.filter(bodega_actual=perfil_bodega)
    elif bodega_filtro:
        base_resumen = base_resumen.filter(bodega_actual_id=bodega_filtro)

    resumen = {
        'total': base_resumen.count(),
        'disponibles': base_resumen.filter(estado='Disponible').count(),
        'asignados': base_resumen.filter(estado='Asignado').count(),
        'mantenimiento': base_resumen.filter(estado='Mantenimiento').count(),
        'baja': base_resumen.filter(estado='Baja').count(),
    }

    if es_bodega(user):
        if perfil_bodega.id == 60:  
            bodegas = PerfilUsuario.objects.filter(rol__nombre='Bodega', estado='Activo').order_by('user__first_name')
        else:
            bodegas = PerfilUsuario.objects.filter(id=perfil_bodega.id)
    else:
        bodegas = PerfilUsuario.objects.filter(rol__nombre='Bodega', estado='Activo').order_by('user__first_name')

    categorias = Categoria.objects.filter(estado='Activo').order_by('nombre')

    if exportar == 'excel':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Bodega"
        ws.append(['ID', 'Elemento', 'Marca', 'Modelo', 'Serial', 'Categoría',
                   'Bodega', 'Estado', 'Fecha Compra', 'Garantía'])
        for e in elementos:
            ws.append([e.id, e.nombre, e.marca, e.modelo, e.serial or '-',
                       e.categoria.nombre,
                       e.bodega_actual.nombre if e.bodega_actual else '-',
                       e.estado,
                       str(e.fecha_compra) if e.fecha_compra else '-',
                       str(e.garantia_hasta) if e.garantia_hasta else '-'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_bodega.xlsx"'
        wb.save(response)
        return response

    if exportar == 'pdf':
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        import io
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elems = [Paragraph("Reporte por Bodega", styles['Title']), Spacer(1, 12)]
        data = [['ID', 'Elemento', 'Marca/Modelo', 'Serial', 'Categoría', 'Bodega', 'Estado', 'Fecha Compra', 'Garantía']]
        for e in elementos:
            data.append([e.id, e.nombre, f"{e.marca} {e.modelo}", e.serial or '-',
                         e.categoria.nombre,
                         e.bodega_actual.nombre if e.bodega_actual else '-',
                         e.estado,
                         str(e.fecha_compra) if e.fecha_compra else '-',
                         str(e.garantia_hasta) if e.garantia_hasta else '-'])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elems.append(t)
        doc.build(elems)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_bodega.pdf"'
        return response

    paginator = Paginator(elementos, 10)
    page = request.GET.get('page', 1)
    try:
        elementos_paginados = paginator.page(page)
    except:
        elementos_paginados = paginator.page(1)

    return render(request, 'apps/inventario/reportes/reporte_bodega_lista.html', {
        'elementos': elementos_paginados,
        'bodegas': bodegas,
        'categorias': categorias,
        'nombres_elementos': NombreElemento.objects.filter(activo=True).order_by('nombre'),
        'resumen': resumen,
        'bodega_filtro': bodega_filtro,
        'estado_filtro': estado_filtro,
        'categoria_filtro': categoria_filtro,
        'nombre_filtro': nombre_filtro,
        'fecha_compra_desde': fecha_compra_desde,
        'fecha_compra_hasta': fecha_compra_hasta,
        'id_filtro':id_filtro,
    })

@login_required_custom
@colaborador_restringido
def reporte_por_usuario(request):
    from apps.usuarios.models import Rol
    from django.http import HttpResponse
    user = request.user
    exportar = request.GET.get('exportar')
    usuario_id_filtro = request.GET.get('usuario_id', '')
    rol_filtro = request.GET.get('rol', '')

    if es_bodega(user):
        usuarios_qs = PerfilUsuario.objects.filter(id=user.perfil.id)
    else:
        usuarios_qs = PerfilUsuario.objects.filter(
            elementos_asignados__isnull=False
        ).distinct().select_related('rol').prefetch_related('elementos_asignados__categoria')

    if usuario_id_filtro:
        usuarios_qs = usuarios_qs.filter(id=usuario_id_filtro)
    if rol_filtro:
        usuarios_qs = usuarios_qs.filter(rol__nombre=rol_filtro)

    roles = Rol.objects.exclude(nombre='Bodega')
    usuarios_lista = PerfilUsuario.objects.filter(
        estado='Activo'
    ).exclude(rol__nombre='Bodega').select_related('rol').order_by('user__first_name')

    if exportar == 'excel':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte por Usuario"
        ws.append(['Usuario', 'Rol', 'ID', 'Elemento', 'Marca', 'Modelo', 'Categoría', 'Estado', 'Fecha Asignación'])
        for u in usuarios_qs:
            for e in u.elementos_asignados.all():
                ws.append([u.nombre, u.rol.nombre if u.rol else '-', e.id, e.nombre,
                           e.marca, e.modelo, e.categoria.nombre, e.estado,
                           str(e.fecha_asignacion) if e.fecha_asignacion else '-'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_usuarios.xlsx"'
        wb.save(response)
        return response

    if exportar == 'pdf':
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        import io
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elems = [Paragraph("Reporte por Usuario", styles['Title']), Spacer(1, 12)]
        for u in usuarios_qs:
            elems.append(Paragraph(f"{u.nombre} ({u.rol.nombre if u.rol else '-'})", styles['Heading2']))
            data = [['ID', 'Elemento', 'Marca/Modelo', 'Categoría', 'Estado', 'Fecha Asignación']]
            for e in u.elementos_asignados.all():
                data.append([e.id, e.nombre, f"{e.marca} {e.modelo}", e.categoria.nombre, e.estado,
                             str(e.fecha_asignacion) if e.fecha_asignacion else '-'])
            t = Table(data, repeatRows=1)
            t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),
                ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
                ('GRID',(0,0),(-1,-1),0.5,colors.black),('FONTSIZE',(0,0),(-1,-1),8)]))
            elems.append(t)
            elems.append(Spacer(1, 12))
        doc.build(elems)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_usuarios.pdf"'
        return response

    paginator = Paginator(usuarios_qs, 10)
    page = request.GET.get('page', 1)
    try:
        usuarios_paginados = paginator.page(page)
    except:
        usuarios_paginados = paginator.page(1)

    return render(request, 'apps/inventario/reportes/reporte_usuarios.html', {
        'usuarios': usuarios_paginados,
        'roles': roles,
        'usuarios_lista': usuarios_lista,
        'usuario_id_filtro': usuario_id_filtro,
        'rol_filtro': rol_filtro,
    })

@login_required_custom
@colaborador_restringido
def reporte_por_categoria(request):
    from django.http import HttpResponse
    exportar = request.GET.get('exportar')
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', '')
    nombre_filtro = request.GET.get('nombre', '')
    fecha_compra_desde = request.GET.get('fecha_desde', '')
    fecha_compra_hasta = request.GET.get('fecha_hasta', '')

    elementos = Elemento.objects.select_related('categoria', 'usuario_actual', 'bodega_actual')

    user = request.user
    if es_bodega(user):
        perfil_bodega = user.perfil
        bodega_obj = PerfilUsuario.objects.filter(id=perfil_bodega.id).first()
        
        if bodega_obj and hasattr(bodega_obj, 'bodega_principal') and bodega_obj.bodega_principal:
            pass
        else:
            elementos = elementos.filter(bodega_actual=perfil_bodega)

    if categoria_filtro:
        elementos = elementos.filter(categoria_id=categoria_filtro)
    if estado_filtro:
        elementos = elementos.filter(estado=estado_filtro)
    if nombre_filtro:
        elementos = elementos.filter(nombre_elemento_id=nombre_filtro)
    if fecha_compra_desde:
        elementos = elementos.filter(fecha_compra__gte=fecha_compra_desde)
    if fecha_compra_hasta:
        elementos = elementos.filter(fecha_compra__lte=fecha_compra_hasta)

    categorias_lista = Categoria.objects.filter(estado='Activo').order_by('nombre')

    if exportar == 'excel':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte por Categoría"
        ws.append(['Categoría', 'ID', 'Elemento', 'Marca', 'Modelo', 'Serial', 'Estado', 'Fecha Compra', 'Usuario Actual', 'Bodega'])
        for e in elementos:
            ws.append([e.categoria.nombre, e.id, e.nombre, e.marca, e.modelo, e.serial or '-',
                       e.estado, str(e.fecha_compra) if e.fecha_compra else '-',
                       e.usuario_actual.nombre if e.usuario_actual else '-',
                       e.bodega_actual.nombre if e.bodega_actual else '-'])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_categoria.xlsx"'
        wb.save(response)
        return response

    if exportar == 'pdf':
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        import io
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elems = [Paragraph("Reporte por Categoría", styles['Title']), Spacer(1, 12)]
        data = [['Categoría', 'ID', 'Elemento', 'Marca/Modelo', 'Serial', 'Estado', 'Fecha Compra', 'Usuario', 'Bodega']]
        for e in elementos:
            data.append([e.categoria.nombre, e.id, e.nombre, f"{e.marca} {e.modelo}",
                         e.serial or '-', e.estado,
                         str(e.fecha_compra) if e.fecha_compra else '-',
                         e.usuario_actual.nombre if e.usuario_actual else '-',
                         e.bodega_actual.nombre if e.bodega_actual else '-'])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('GRID',(0,0),(-1,-1),0.5,colors.black),('FONTSIZE',(0,0),(-1,-1),8)]))
        elems.append(t)
        doc.build(elems)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_categoria.pdf"'
        return response

    paginator = Paginator(elementos, 10)
    page = request.GET.get('page', 1)
    try:
        elementos_paginados = paginator.page(page)
    except:
        elementos_paginados = paginator.page(1)

    return render(request, 'apps/inventario/reportes/reporte_categoria_nuevo.html', {
        'elementos': elementos_paginados,
        'categorias_lista': categorias_lista,
        'nombres_elementos': NombreElemento.objects.filter(activo=True).order_by('nombre'),
        'categoria_filtro': categoria_filtro,
        'estado_filtro': estado_filtro,
        'nombre_filtro': nombre_filtro,
        'fecha_compra_desde': fecha_compra_desde,
        'fecha_compra_hasta': fecha_compra_hasta,
    })

from apps.inventario.acta_movimiento import generar_acta_movimiento

@login_required_custom
@colaborador_restringido
def reporte_garantias(request):
    from django.http import HttpResponse
    from datetime import datetime, timedelta
    
    exportar = request.GET.get('exportar')
    categoria_filtro = request.GET.get('categoria', '')
    estado_filtro = request.GET.get('estado', '')
    nombre_filtro = request.GET.get('nombre', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    # Solo elementos que tengan fecha de garantía registrada
    elementos = Elemento.objects.filter(garantia_hasta__isnull=False).select_related('categoria', 'usuario_actual', 'bodega_actual')

    user = request.user
    if es_bodega(user):
        perfil_bodega = user.perfil
        if perfil_bodega.id != 1:  # Asumiendo 1 es la bodega principal general
            elementos = elementos.filter(bodega_actual=perfil_bodega)

    if categoria_filtro:
        elementos = elementos.filter(categoria_id=categoria_filtro)
    if estado_filtro:
        elementos = elementos.filter(estado=estado_filtro)
    if nombre_filtro:
        elementos = elementos.filter(nombre_elemento_id=nombre_filtro)
    if fecha_desde:
        elementos = elementos.filter(garantia_hasta__gte=fecha_desde)
    if fecha_hasta:
        elementos = elementos.filter(garantia_hasta__lte=fecha_hasta)
    vencimiento_filtro = request.GET.get('vencimiento', '')
    categoria_filtro = request.GET.get('categoria', '')
    # ... otros filtros ...

    elementos = Elemento.objects.filter(garantia_hasta__isnull=False).select_related('categoria', 'usuario_actual', 'bodega_actual')        

    elementos = elementos.order_by('garantia_hasta') # Ordenar por los que vencen primero

    categorias_lista = Categoria.objects.filter(estado='Activo').order_by('nombre')
    nombres_elementos = NombreElemento.objects.filter(activo=True).order_by('nombre')

    # Cálculos para el resumen
    hoy = datetime.now().date()
    proximos_30_dias = hoy + timedelta(days=30)

    if vencimiento_filtro == 'vencida':
        elementos = elementos.filter(garantia_hasta__lt=hoy)
    elif vencimiento_filtro == 'vigente':
        elementos = elementos.filter(garantia_hasta__gte=hoy)
    elif vencimiento_filtro == 'proxima':
        elementos = elementos.filter(garantia_hasta__gte=hoy, garantia_hasta__lte=proximos_30_dias)

    resumen = {
        'total': elementos.count(),
        'vigentes': elementos.filter(garantia_hasta__gte=hoy).count(),
        'vencidas': elementos.filter(garantia_hasta__lt=hoy).count(),
        'proximas': elementos.filter(garantia_hasta__gte=hoy, garantia_hasta__lte=proximos_30_dias).count(),
    }

    if exportar == 'excel':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte Garantias"
        ws.append(['ID', 'Elemento', 'Marca', 'Modelo', 'Serial', 'Categoría', 'Ubicación', 'Vencimiento Garantía'])
        for e in elementos:
            ubicacion = e.bodega_actual.nombre if e.bodega_actual else (e.usuario_actual.nombre if e.usuario_actual else 'Baja/Sin asignar')
            ws.append([e.id, e.nombre, e.marca, e.modelo, e.serial or '-', e.categoria.nombre, ubicacion, str(e.garantia_hasta)])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_garantias.xlsx"'
        wb.save(response)
        return response

    if exportar == 'pdf':
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        import io
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elems = [Paragraph("Reporte de Garantías", styles['Title']), Spacer(1, 12)]
        data = [['ID', 'Elemento', 'S/N', 'Categoría', 'Ubicación', 'Vencimiento']]
        for e in elementos:
            ubicacion = e.bodega_actual.nombre if e.bodega_actual else (e.usuario_actual.nombre if e.usuario_actual else '-')
            data.append([e.id, e.nombre, e.serial or '-', e.categoria.nombre, ubicacion, str(e.garantia_hasta)])
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.grey),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),('GRID',(0,0),(-1,-1),0.5,colors.black),('FONTSIZE',(0,0),(-1,-1),8)]))
        elems.append(t)
        doc.build(elems)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_garantias.pdf"'
        return response

    paginator = Paginator(elementos, 10)
    page = request.GET.get('page', 1)
    try:
        elementos_paginados = paginator.page(page)
    except:
        elementos_paginados = paginator.page(1)

    return render(request, 'apps/inventario/reportes/reporte_garantias.html', {
        'elementos': elementos_paginados,
        'categorias_lista': categorias_lista,
        'nombres_elementos': nombres_elementos,
        'categoria_filtro': categoria_filtro,
        'estado_filtro': estado_filtro,
        'nombre_filtro': nombre_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'resumen': resumen,
        'vencimiento_filtro': vencimiento_filtro,
    })

@login_required_custom
def descargar_acta(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logocorp.png') # o staticfiles si usas collectstatic
    logo_path=logo_path.replace('\\', '/')  # Asegurarse de que las barras sean correctas para el sistema operativo

    # 1. Obtenemos la plantilla HTML
    template = get_template('apps/inventario/movimientos/acta_pdf.html')
    
    # 2. Pasamos los datos del movimiento al HTML
    context = {'movimiento': movimiento,
               'logo_path': logo_path,}
    html = template.render(context)
    
    # 3. Creamos la respuesta PDF y la forzamos a descargar (attachment)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="acta_MOV-{movimiento_id}.pdf"'
    
    # 4. Generamos el PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Tuvimos algunos errores <pre>' + html + '</pre>')
    return response

@login_required_custom
def preview_acta(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logocorp.png') # o staticfiles si usas collectstatic
    logo_path=logo_path.replace('\\', '/')  # Asegurarse de que las barras sean correctas para el sistema operativo
    
    # 1. Obtenemos la plantilla HTML
    template = get_template('apps/inventario/movimientos/acta_pdf.html')
    
    # 2. Pasamos los datos del movimiento al HTML
    context = {'movimiento': movimiento,
               'logo_path': logo_path,}
    html = template.render(context)
    
    # 3. Creamos la respuesta PDF y la forzamos a mostrar en el navegador (inline)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="acta_MOV-{movimiento_id}.pdf"'
    
    # 4. Generamos el PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Tuvimos algunos errores <pre>' + html + '</pre>')
    return response

@login_required_custom
def firmar_acta(request, movimiento_id, tipo_firma):
    movimiento = get_object_or_404(MovimientoInventario, id=movimiento_id)
    perfil = request.user.perfil

    if movimiento.es_firmado_total:
        messages.error(request, 'Este documento ya fue firmado completamente y no puede modificarse.')
        return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)

    if tipo_firma == 'recibe':
        if movimiento.usuario_destino != perfil:
            messages.error(request, 'Solo el usuario destinatario puede firmar como "Recibe".')
            return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)
        if movimiento.firma_recibe:
            messages.error(request, 'Esta firma ya fue registrada.')
            return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)

    elif tipo_firma == 'elabora':
        if movimiento.usuario_registro != perfil:
            messages.error(request, 'Solo quien registró el movimiento puede firmar como "Elabora".')
            return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)
        if movimiento.firma_elabora:
            messages.error(request, 'Esta firma ya fue registrada.')
            return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)

    elif tipo_firma == 'autoriza':
        if not movimiento.usuario_autoriza:
            messages.error(request, 'Este movimiento no tiene asignado un usuario autorizador.')
            return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)
            
        if movimiento.usuario_autoriza != perfil:
            messages.error(
                request, 
                f'Acceso denegado. Solo {movimiento.usuario_autoriza.nombre} fue asignado para firmar como "Autoriza".'
            )
            return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)
        
        if movimiento.firma_autoriza:
            messages.error(request, 'Esta firma ya fue registrada.')
            return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)
    else:
        messages.error(request, 'Tipo de firma no válido.')
        return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)

    if request.method == 'POST':
        firma_data = request.POST.get('firma_data', '')
        if not firma_data or 'base64' not in firma_data:
            messages.error(request, 'No se recibió la firma. Intente de nuevo.')
            return redirect('inventario:firmar_acta', movimiento_id=movimiento_id, tipo_firma=tipo_firma)

        try:
            # 1. Guardamos la firma usando el servicio
            guardar_firma_acta(movimiento, tipo_firma, firma_data)
            
            # 2. Refrescamos desde la Base de Datos para verificar el estado
            movimiento.refresh_from_db()
            
            # 3. Si ya se completaron las firmas, generamos y guardamos el PDF
            if movimiento.es_firmado_total and not movimiento.documento_soporte:
                movimiento.plicar_movimiento_a_elementos()
                buffer = generar_acta_movimiento(movimiento)
                
                from django.core.files.base import ContentFile
                pdf_bytes = buffer.getvalue() 
                nombre_pdf = f"acta_movimiento_{movimiento.id}_firmada.pdf"
                
                movimiento.documento_soporte.save(nombre_pdf, ContentFile(pdf_bytes), save=True)

            etiquetas_msg = {'recibe': 'Recibe', 'elabora': 'Elaborado por', 'autoriza': 'Autoriza'}
            messages.success(request, f'Firma "{etiquetas_msg.get(tipo_firma, tipo_firma)}" registrada correctamente.')
            return redirect('inventario:movimiento_detalle', movimiento_id=movimiento_id)

        except Exception as e:
            messages.error(request, f'Error al procesar la firma: {str(e)}')

    etiquetas = {
        'recibe': 'Recibe',
        'elabora': 'Elaborado por',
        'autoriza': 'Autoriza',
    }

    return render(request, 'apps/inventario/movimientos/firmar_acta.html', {
        'movimiento': movimiento,
        'tipo_firma': tipo_firma,
        'etiqueta_firma': etiquetas.get(tipo_firma, tipo_firma),
    })

@login_required_custom
def mis_documentos(request):
    """Vista para usuarios no técnicos: sus movimientos pendientes de firma."""
    perfil = request.user.perfil

    # Movimientos donde el usuario está involucrado
    if es_colaborador(request.user):
        movimientos = MovimientoInventario.objects.filter(
            Q(usuario_destino=perfil) | Q(usuario_autoriza=perfil)
        ).select_related(
            'usuario_origen', 'usuario_destino', 'usuario_registro'
        ).order_by('-fecha_movimiento')
    else:
        movimientos = MovimientoInventario.objects.filter(
            Q(usuario_destino=perfil) |
            Q(usuario_origen=perfil) |
            Q(usuario_registro=perfil)|
            Q(usuario_autoriza=perfil)
        ).select_related(
            'usuario_origen', 'usuario_destino', 'usuario_registro'
        ).order_by('-fecha_movimiento')

    # Filtrar pendientes
    # 1. Separamos los pendientes en tres "cajones" distintos
    pendientes_recibir = movimientos.filter(
        Q(usuario_destino=perfil) & (Q(firma_recibe='') | Q(firma_recibe__isnull=True))
    ).distinct()

    pendientes_autorizar = movimientos.filter(
        Q(usuario_autoriza=perfil) & (Q(firma_autoriza='') | Q(firma_autoriza__isnull=True))
    ).distinct()

    pendientes_elaborar = movimientos.filter(
        Q(usuario_registro=perfil) & (Q(firma_elabora='') | Q(firma_elabora__isnull=True))
    ).distinct()

    # 2. Obtenemos los firmados excluyendo todos los que aún están en los 3 cajones de pendientes
    todos_pendientes_ids = list(pendientes_recibir.values_list('id', flat=True)) + \
                           list(pendientes_autorizar.values_list('id', flat=True)) + \
                           list(pendientes_elaborar.values_list('id', flat=True))
    
    firmados_qs = movimientos.exclude(id__in=todos_pendientes_ids)

    # 3. Paginación: Le damos a cada lista su propio paginador
    recibir_paginados = Paginator(pendientes_recibir, 10).get_page(request.GET.get('page_r', 1))
    autorizar_paginados = Paginator(pendientes_autorizar, 10).get_page(request.GET.get('page_a', 1))
    elaborar_paginados = Paginator(pendientes_elaborar, 10).get_page(request.GET.get('page_e', 1))
    firmados_paginados = Paginator(firmados_qs, 10).get_page(request.GET.get('page_f', 1))

    # ✅ Pasamos las TRES nuevas listas al contexto del HTML
    return render(request, 'apps/inventario/movimientos/mis_documentos.html', {
        'pendientes_recibir': recibir_paginados,
        'pendientes_autorizar': autorizar_paginados,
        'pendientes_elaborar': elaborar_paginados,
        'firmados': firmados_paginados,
        'perfil': perfil,
    })
@login_required_custom
def mis_elementos_asignados(request):
    """Vista dedicada para que técnicos/bodega vean los elementos que tienen asignados o a su cargo."""
    perfil = request.user.perfil
    
    # --- LÓGICA DE FILTRADO SEGÚN ROL ---
    if perfil.rol.nombre == 'Bodega':
        # Si el usuario es de rol Bodega, filtramos por la columna 'bodega_actual'
        # Esto traerá todos los equipos que están físicamente en esa bodega (Disponibles, etc.)
        elementos = Elemento.objects.filter(
            bodega_actual=perfil
        )
    else:
        # Para Técnicos o Colaboradores, seguimos filtrando por 'usuario_actual'
        # Esto trae los equipos que tienen bajo su nombre (Asignados)
        elementos = Elemento.objects.filter(
            usuario_actual=perfil
        )
    
    # Ordenamos y optimizamos la consulta (mantenemos tu lógica original)
    elementos = elementos.select_related('categoria', 'bodega_actual', 'nombre_elemento').order_by('-fecha_registro')
    
    # --- PAGINACIÓN (Mantener consistencia) ---
    paginator = Paginator(elementos, 10)
    page = request.GET.get('page', 1)
    
    try:
        elementos_paginados = paginator.page(page)
    except PageNotAnInteger:
        elementos_paginados = paginator.page(1)
    except EmptyPage:
        elementos_paginados = paginator.page(paginator.num_pages)
        
    # Consultas extra para los filtros de la plantilla 'lista.html'
    categorias = Categoria.objects.filter(estado='Activo').order_by('nombre')
    bodegas = PerfilUsuario.objects.filter(rol__nombre='Bodega', estado='Activo').order_by('user__first_name')
    
    context = {
        'elementos': elementos_paginados,
        'categorias': categorias,
        'bodegas': bodegas,
        'es_mis_elementos': True, 
        'titulo_personalizado': "Mi Inventario a Cargo" # Opcional: para cambiar el título en el HTML
    }
    
    return render(request, 'apps/inventario/elementos/lista.html', context)

@login_required_custom
def api_detalle_elemento(request, elemento_id):
    """Retorna los datos clave de un elemento para la automatización del formulario."""
    try:
        elemento = Elemento.objects.get(id=elemento_id)
        return JsonResponse({
            'fecha_compra': elemento.fecha_compra.strftime('%Y-%m-%d') if elemento.fecha_compra else '',
            'garantia_hasta': elemento.garantia_hasta.strftime('%Y-%m-%d') if elemento.garantia_hasta else '',
        })
    except Elemento.DoesNotExist:
        return JsonResponse({'error': 'No encontrado'}, status=404)
    
def mis_elementos(request):
    # Traemos los elementos donde el "usuario_actual" es el perfil que está logueado
    elementos = Elemento.objects.filter(usuario_actual=request.user.perfil)
    
    return render(request, 'apps/inventario/elementos/mis_elementos.html', {
        'elementos': elementos
    })    

        